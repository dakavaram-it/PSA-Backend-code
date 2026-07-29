"""Phase 1 of the staged IVRS load: read the wave Excel into a CONTENTION-FREE
staging table we own (dakavara_pa.pulse_trend_ivrs31_stg). No auto-increment, no
secondary indexes → fast inserts, no lock contention with the live master table.

Phase 2 (separate, scripts/merge_ivrs.py) moves staging → survey24.ivrs_survey_answer
with a single INSERT … SELECT.

Usage:
    PYTHONPATH=. python3.13 scripts/stage_ivrs.py "<file.xlsx>" [--sid 31] [--date 2026-06-04] [--batch 5000]
"""
import sys
import time
from collections import Counter

import openpyxl
from sqlalchemy import text

from app.database.db import dakavara_session

STAGE = "dakavara_pa.pulse_trend_ivrs31_stg"
COL = {"mobile": 0, "acid": 2, "sid": 5, "round": 6, "clip": 7, "option_no": 8, "qid": 9, "opt_id": 10}
DATA_SHEETS = ("SHEET_1", "SHEET_2")
DDL = f"""
CREATE TABLE IF NOT EXISTS {STAGE} (
  mobile_no VARCHAR(15), constituency_id INT, ivrs_survey_id INT, round_id INT,
  clip_no VARCHAR(20), option_no INT, ivrs_question_id INT, ivrs_option_id INT,
  survey_date DATE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""
INS = text(f"""INSERT INTO {STAGE}
  (mobile_no, constituency_id, ivrs_survey_id, round_id, clip_no, option_no, ivrs_question_id, ivrs_option_id, survey_date)
  VALUES (:mobile_no,:constituency_id,:ivrs_survey_id,:round_id,:clip_no,:option_no,:ivrs_question_id,:ivrs_option_id,:survey_date)""")


def arg(flag, d):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def _int(v):
    return int(v) if v not in (None, "") else None


def main():
    path = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
    sid = int(arg("--sid", "31")); survey_date = arg("--date", "2026-06-04"); batch = int(arg("--batch", "5000"))
    t0 = time.time()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in DATA_SHEETS if s in wb.sheetnames] or [s for s in wb.sheetnames if s.lower().startswith("sheet")]
    rows, opt = [], Counter()
    for sn in sheets:
        for i, row in enumerate(wb[sn].iter_rows(values_only=True)):
            if i == 0:
                continue
            if _int(row[COL["sid"]]) != sid or _int(row[COL["opt_id"]]) is None or row[COL["mobile"]] in (None, ""):
                continue
            opt[_int(row[COL["opt_id"]])] += 1
            rows.append({"mobile_no": str(row[COL["mobile"]]).strip(), "constituency_id": _int(row[COL["acid"]]),
                         "ivrs_survey_id": sid, "round_id": _int(row[COL["round"]]),
                         "clip_no": (str(row[COL["clip"]]) if row[COL["clip"]] is not None else None),
                         "option_no": _int(row[COL["option_no"]]), "ivrs_question_id": _int(row[COL["qid"]]),
                         "ivrs_option_id": _int(row[COL["opt_id"]]), "survey_date": survey_date})
    print(f"read {len(rows):,} rows in {(time.time()-t0)/60:.1f} min; option dist {dict(opt)}", flush=True)

    t1 = time.time(); n = 0
    with dakavara_session() as db:
        db.execute(text(DDL)); db.execute(text(f"TRUNCATE TABLE {STAGE}")); db.commit()
        for j in range(0, len(rows), batch):
            db.execute(INS, rows[j:j + batch]); db.commit(); n += len(rows[j:j + batch])
            if n % 100000 < batch:
                print(f"  staged {n:,}/{len(rows):,} ({(time.time()-t1)/60:.1f} min)", flush=True)
    print(f"STAGED {n:,} rows into {STAGE} in {(time.time()-t1)/60:.1f} min", flush=True)


if __name__ == "__main__":
    main()
