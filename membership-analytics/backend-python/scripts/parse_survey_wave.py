"""Parse a RAW vendor survey workbook (Dhruva / Codemo / CSDS) into survey24's
long-format answer rows — the transform the existing loaders assume already happened.

Pipeline position
-----------------
  raw vendor .xlsx  ──(this script)──►  long rows {mobile, acid, qid, option_id, ...}
        │                                          │
        │                                  --commit │
        ▼                                          ▼
  DRY-RUN report                      dakavara_pa.<stage>  ──merge_ivrs.py──►  survey24.ivrs_survey_answer

It does NOT touch the live 34M-row master directly. With --commit it writes a staging
table we own (same contention-free pattern as stage_ivrs.py); the Phase-2 merge stays unchanged.

Design (grounded in how sid 31 — June'26 IVRS — is actually stored)
-------------------------------------------------------------------
* sid 31 stores ONLY the voting-intention question (q10: NDA=28/YSRCP=15/OTHERS=16).
  Caste/age/gender are NOT answer rows — they come from survey24.ivrs_mobiles via mobile_no.
  → so caste & category are intentionally NOT ingested as answers here.
* Option codes are QUESTION-SCOPED: raw "TDP" = option 14 in a vote question but 47
  ("Anyone from TDP") in the party-visited question; "Others" = 16 for vote but no option
  for issues/CM. SCOPES below encode that; a shared SENTINELS set drops non-responses.
* The party/option ids (14/15/16/28 …) are the SAME ones pulse_config.py depends on —
  see survey24_touchpoints.md §A. Do not renumber them.

Usage
-----
  PYTHONPATH=. python3.13 scripts/parse_survey_wave.py "<raw.xlsx>" --vendor codemo --sid 41 --date 2026-06-01
  PYTHONPATH=. python3.13 scripts/parse_survey_wave.py "<raw.xlsx>" --vendor dhruva --sid 31 --date 2026-05-05 --commit
Flags: --vendor {dhruva|codemo|csds}  --sid N  --date YYYY-MM-DD  --batch 5000  --commit
DRY-RUN (default) prints row/option/ignored distributions AND any UNMAPPED value so gaps surface.
"""
import re
import sys
import time
from collections import Counter, defaultdict

import openpyxl
from sqlalchemy import text

from app.database.db import dakavara_session

# ── survey24.ivrs_question ids ──
Q_VOTE_ASSEMBLY, Q_VOTE_ALLIANCE, Q_VOTE_PAST_2024 = 10, 11, 8
Q_CM_1ST, Q_CM_2ND = 12, 13
Q_GOVT_OPINION, Q_MLA_OPINION, Q_MLA_AVAILABLE = 14, 15, 16
Q_MAIN_PROBLEM, Q_PARTY_VISITED = 17, 18
Q_RELIGION, Q_EDUCATION, Q_OCCUPATION = 19, 22, 23

# Non-responses → ignored (not counted, not "unmapped"), across every question.
SENTINELS = {
    "na", "n/a", "", "none", "not sure", "do not know", "do not know mla", "dont know",
    "phone disconnected", "phone switched off", "switched off", "phone busy",
    "do not want to answer", "no response", "ringing no response", "call not connected",
    "undecided", "do not remember", "dont remember",
    "don't know / can't say", "don’t know / can’t say", "dont know / cant say",
    "don't know/can't say", "don’t know/can’t say", "can't say", "cant say",
    "no opinion", "not applicable", "na / not applicable", "did not apply / na",
}

# Per-question value→option_id overrides for the AMBIGUOUS short tokens. Everything not
# listed falls back to the live ivrs_option NAME match (load_codebook) — which already
# covers the long tail (issue names, occupations, education, full CM names, ratings).
# A value mapped to None is explicitly ignored (real answer, but no representative option).
SCOPES = {
    "vote": {  # party / alliance / past-vote. Minor parties/independents → 16 (Others),
               # matching the published NDA/YSRCP/Others bucketing the dashboard counts.
        "tdp": 14, "ysrcp": 15, "ysrc": 15, "others": 16, "other": 16,
        "nda": 28, "janasena": 29, "jsp": 29, "bjp": 30, "inc": 22, "bsp": 23,
        "did not vote": 44, "didnt vote": 44, "nota": 46, "india": 45,
        "others/independent": 16, "independent": 16, "ind": 16,
        "cpi": 16, "cpi(m)": 16, "cpi (m)": 16, "brs": 16, "aifb": 16, "trs": 16,
        "sdpi": 16, "aimim": 16, "mim": 16, "jnp": 29,
        "rpia": 16, "jbnp": 16, "bcyp": 16, "aijp": 16, "arps": 16, "nacp": 16, "bap": 16,
        # CSDS spells parties in full — must map (esp. YSRCP) before the Others catch-all
        "telugu desam party (tdp)": 14, "jana sena party (jsp)": 29,
        "yuvajana sramika rythu congress party (ysr congress)": 15,
        "ysr congress party": 15, "ysr congress": 15, "ysrc party": 15,
        "bharatiya janata party (bjp)": 30, "indian national congress (inc)": 22,
        "ysr congress party (ysr congress)": 15, "other": 16, "nota": 46,
    },
    "cm": {  # chief-minister choice; full names match ivrs_option via base map
        "cbn": 33, "chandrababu": 33, "chandrababu naidu": 33,
        "nara chandra babu naidu": 33, "chandra babu naidu": 33, "nara chandrababu naidu": 33,
        "jagan": 34, "ys jagan": 34, "jagan mohan reddy": 34, "ys jagan mohan reddy": 34,
        "lokesh": 35, "pawan": 36, "konidela pawan kalyan": 36,
        "sharmila": 37, "y. s. sharmila reddy": 37, "y.s. sharmila reddy": 37,
        "others": None, "other": None,
    },
    "rating": {},   # very good/good/average/bad/very bad → base map
    "avail":  {},   # fairly/somewhat/highly/hardly/not available at all → base map
    "issue": {      # issue names → base map; only fix variants + drop catch-all "others"
        "no issue": 57, "increased electricity bills": 65, "capital development": 79,
        "others": None, "other": None,
        # CSDS Q4 short labels
        "roads": 58, "drainage": 61, "job": 59, "jobs": 59, "water": 60,
        "price rise/inflation": 72, "inflation": 72, "corruption": 71, "health": 73,
        "education": 69, "electricity": 67, "transport": 75, "agriculture": 63,
    },
    "visited": {    # which party's workers visited; "Anyone from X" options
        "tdp": 47, "bjp": 48, "inc": 49, "ysrcp": 50,
        "jsp": None, "janasena": None, "bsp": None, "cpi": None, "ind": None,
        "other party": None, "others": None,  # no representative option
    },
    "religion": {   # hindu/christian/muslim/buddhist → base map; variants below
        "boudh": 88,            # Buddhist
        "sikh": None, "jain": None,  # no option today — ignored (seed options to capture)
    },
    "base": {},     # education / occupation — name match only
    "satis": {      # Dhruva satisfaction collapsed onto the existing opinion scale.
                    # Dhruva phrasing is "Unsatisfied" (not "Dissatisfied") — both handled.
        "completely satisfied": 39, "somewhat satisfied": 40, "neutral": 41,
        "somewhat dissatisfied": 42, "completely dissatisfied": 43,
        "somewhat unsatisfied": 42, "completely unsatisfied": 43,
    },
    "devscale": {   # development perception (Improved / Deteriorated / No change)
        "improved": 129, "deteriorated": 130,
        "no change": 131, "no change / it has stayed the same": 131,
        # June'25 used a BINARY Yes/No instrument here — not comparable to the 3-point
        # ordinal, so we drop it rather than mis-map No→Deteriorated.
        "yes": None, "no": None,
    },
    "agree": {      # MLA/MP attribute battery (Completely/Somewhat Agree/Disagree)
        "completely agree": 132, "somewhat agree": 133,
        "somewhat disagree": 134, "completely disagree": 135,
    },
}

# ── AC name reconciliation: vendor spelling → survey24.ivrs_mobiles.constituency_name ──
AC_ALIASES = {
    "bhimli": "bhimili", "cheepurupalle": "cheepurupalli", "denduluru": "dendulur",
    "gopalapuram": "gopalpuram", "gurazala": "gurzala", "ichchapuram": "ichapuram",
    "jaggayyapeta": "jaggayyapet", "kondapi": "kondepi", "madanapalle": "madanpalle",
    "markapuram": "markapur", "narasapuram": "narasapur", "pedakurapadu": "peddakurapadu",
    "prathipadu (sc)": "prathipadu", "rajamundry rural": "rajahmundry rural",
    "rayachoti": "rayachoty", "sullurpeta": "sullurpet", "tadikonda (sc)": "tadikonda",
    "tadipatri": "tadpatri", "unguturu": "ungutur", "v.madugula": "madugula",
    "vemuru (sc)": "vemuru", "vijaywada west": "vijayawada west",
    # Dhruva spelling variants
    "sattenapalle": "sattenapalli", "ponnuru": "ponnur", "pulivendula": "pulivendla",
    # CSDS spelling variants
    "gurajala": "gurzala", "narsapuram": "narasapur", "payakaraopet": "payakaraopeta",
    "yelamanchili": "elamanchili", "chipurupalle": "cheepurupalli",
    "ramchandrapuram": "ramachandrapuram", "anakapalle": "anakapalli",
    "rampachodovaram": "rampachodavaram",
    # Dhruva has no district column to split Gannavaram → default bare to Krishna, (SC) to EG
    "gannavaram": "gannavaram (krishna)", "gannavaram (sc)": "gannavaram (eg)",
}
AC_DISTRICT_ALIASES = {  # one vendor name → different seats by district
    ("gannavaram", "krishna"): "gannavaram (krishna)",
    ("gannavaram", "konaseema"): "gannavaram (eg)",
    ("gannavaram (sc)", "konaseema"): "gannavaram (eg)",
}


def _norm(v):
    return re.sub(r"\s+", " ", str(v).strip().lower()) if v not in (None, "") else None


# ── VENDOR_PROFILES: raw column → (ivrs_question_id, scope) ──
VENDOR_PROFILES = {
    "codemo": {
        "id_col": "id", "mobile_col": "phone no", "ac_col": "ac", "ac_kind": "name",
        "district_col": "district",
        "col_aliases": {"mla ratinga": "mla rating"},   # header typo in June'25 wave
        "questions": {
            "party vote upcoming election": (Q_VOTE_ASSEMBLY, "vote"),
            "alliance vote upcoming election": (Q_VOTE_ALLIANCE, "vote"),
            "party voted in ae24": (Q_VOTE_PAST_2024, "vote"),
            "cm 1st choice": (Q_CM_1ST, "cm"), "cm 2nd choice": (Q_CM_2ND, "cm"),
            "govt rating": (Q_GOVT_OPINION, "rating"), "mla rating": (Q_MLA_OPINION, "rating"),
            "mla availability": (Q_MLA_AVAILABLE, "avail"),
            "biggest issue": (Q_MAIN_PROBLEM, "issue"),
            "party_workers_visited_home": (Q_PARTY_VISITED, "visited"),
            "religion": (Q_RELIGION, "religion"),
            "education": (Q_EDUCATION, "base"), "occupation": (Q_OCCUPATION, "base"),
        },
    },
    "dhruva": {
        # Resolve by NAME, not AC_num — verified AC_num(131,Jammalamadugu) ≠ constituency_id(244).
        "id_col": "submission_id", "mobile_col": None, "ac_col": "ac_name", "ac_kind": "name",
        "district_col": None,
        "col_aliases": {                                 # header drift across waves
            "voting intention": "voting_intention_vs",   # June'25 name
            "vs past final": "vs past",                  # June'25 name
            # June'25 names the MLA/MP attribute battery differently → map to canonical cols
            "mla_satisfaction_hardworking_available": "mla_hardworking",
            "mla_satisfaction_accessibility": "mla_availability",
            "mla_satisfaction_honest": "mla_honesty",
            "mla_satisfaction_cares_for_people": "mla_cares_community",
            "mp_satisfaction_hardworking_available": "mp_hardworking",
            "mp_satisfaction_honest": "mp_honesty",
            "mp_satisfaction_development": "mp_work_development",
        },
        "questions": {
            "voting_intention_vs": (Q_VOTE_ASSEMBLY, "vote"), "vs past": (Q_VOTE_PAST_2024, "vote"),
            "ls past": (28, "vote"),
            "cm choice": (Q_CM_1ST, "cm"),
            "state_govt_satisfaction": (Q_GOVT_OPINION, "satis"),
            "mla satisfaction": (Q_MLA_OPINION, "satis"),
            "mla availability": (Q_MLA_AVAILABLE, "avail"),
            "religion": (Q_RELIGION, "religion"),
            # leader satisfaction (reuse opinion scale 39–43)
            "cm satisfaction": (24, "satis"), "dy cm satisfaction": (25, "satis"),
            "mp satisfaction": (26, "satis"), "nara lokesh satisfaction": (27, "satis"),
            # development perception (Improved/Deteriorated/No change)
            "key_area_development_law_order": (29, "devscale"),
            "key_area_development_education": (30, "devscale"),
            "key_area_development_employment": (31, "devscale"),
            "key_area_development_healthcare": (32, "devscale"),
            "key_area_development_electricity_tariff": (33, "devscale"),
            "key_area_development_electricity_availability": (33, "devscale"),
            "key_area_development_welfare_schemes": (34, "devscale"),
            "key_area_development_road_infrastructure": (35, "devscale"),
            "key_area_development_public_transport": (36, "devscale"),
            "key_area_development_investments_to_state": (37, "devscale"),
            "key_area_development_inflation": (38, "devscale"),
            "key_area_development_drainage": (39, "devscale"),
            # MLA/MP attribute battery (Completely/Somewhat Agree/Disagree)
            "mla_hardworking": (40, "agree"), "mla_availability": (41, "agree"),
            "mla_honesty": (42, "agree"), "mla_work_development": (43, "agree"),
            "mla_cares_community": (44, "agree"),
            "mp_hardworking": (45, "agree"), "mp_availability": (46, "agree"),
            "mp_honesty": (47, "agree"), "mp_work_development": (48, "agree"),
        },
    },
    "csds": {  # 109-col coded F2F. Values are self-describing "NN: Label" → strip_code.
        "id_col": "caseid", "mobile_col": "z10", "ac_col": "f3", "ac_kind": "name",
        "district_col": None, "strip_code": True,
        "questions": {
            # development perception battery Q1a–Q1k (same 11 areas, same order as the bank)
            "q1a": (29, "devscale"), "q1b": (30, "devscale"), "q1c": (31, "devscale"),
            "q1d": (32, "devscale"), "q1e": (33, "devscale"), "q1f": (34, "devscale"),
            "q1g": (35, "devscale"), "q1h": (36, "devscale"), "q1i": (37, "devscale"),
            "q1j": (38, "devscale"), "q1k": (39, "devscale"),
            "q3": (14, "rating"),          # rate CBN state govt → govt opinion
            "q4": (17, "issue"),           # major issue
            "q12": (24, "rating"),         # CM (CBN) performance → CM
            "q21": (25, "rating"),         # Dy CM (Pawan) → Dy CM
            "q22": (27, "rating"),         # Nara Lokesh
            "q23": (15, "rating"),         # MLA → MLA opinion
            "q29": (26, "rating"),         # MP → MP
            "q36": (8, "vote"),            # 2024 assembly vote (past)
            "q37": (28, "vote"),           # 2024 Lok Sabha vote (past)
            "q39": (10, "vote"),           # if assembly election today
            "q41": (12, "cm"),             # best next CM
            "z6": (19, "religion"),        # religion
        },
    },
}


def strip_code(v):
    """CSDS values look like '01: Improved' / '81: VIJAYAWADA EAST' — drop the leading code.
    A bare code with no label (e.g. '8', '32') is unusable → return '' so it's ignored."""
    if v in (None, ""):
        return v
    s = re.sub(r"^\s*\d+\s*[:\-]\s*", "", str(v))
    return "" if re.fullmatch(r"\d+", s.strip()) else s


def arg(flag, d=None):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def load_codebook(db):
    """base option-name → id (from live ivrs_option), and AC name → constituency_id."""
    base = {}
    for oid, name in db.execute(text("SELECT ivrs_option_id, option_name FROM survey24.ivrs_option")):
        if _norm(name):
            base[_norm(name)] = oid
    ac_by_name = {}
    for cid, name in db.execute(text(
            "SELECT DISTINCT constituency_id, constituency_name FROM survey24.ivrs_mobiles "
            "WHERE constituency_name IS NOT NULL")):
        if _norm(name):
            ac_by_name[_norm(name)] = cid
    return base, ac_by_name


def resolve_option(raw, scope, base):
    """→ (option_id, status) where status ∈ {'ok','ignore','miss'}."""
    n = _norm(raw)
    if n is None or n in SENTINELS:
        return None, "ignore"
    sc = SCOPES.get(scope, {})
    if n in sc:
        oid = sc[n]
        return (oid, "ok") if oid is not None else (None, "ignore")
    # independents come as unbounded "IND_<candidate name>" — bucket as Others in vote qs
    if scope == "vote" and (n.startswith("ind_") or n.startswith("ind ")):
        return 16, "ok"
    if n in base:
        return base[n], "ok"
    # vote questions have a long tail of minor registered parties — all are "Others" to the
    # dashboard's NDA/YSRCP/Others bucketing (major parties are already mapped explicitly above).
    if scope == "vote":
        return 16, "ok"
    # issue answers are semi-free-text (esp. CSDS Q4): keep the ones with a representative
    # option, drop the long free-text tail rather than chase every phrasing.
    if scope == "issue":
        return None, "ignore"
    return None, "miss"


def header_index(ws):
    for row in ws.iter_rows(values_only=True):
        return {_norm(c): i for i, c in enumerate(row) if c is not None}
    return {}


def main():
    path = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
    vendor = (arg("--vendor") or "").lower()
    sid = int(arg("--sid", "0")); survey_date = arg("--date"); batch = int(arg("--batch", "5000"))
    commit = "--commit" in sys.argv
    if not path or vendor not in VENDOR_PROFILES or not sid or not survey_date:
        print("usage: parse_survey_wave.py <raw.xlsx> --vendor {dhruva|codemo|csds} --sid N --date YYYY-MM-DD [--commit]")
        return
    prof = VENDOR_PROFILES[vendor]
    if not prof["questions"]:
        print(f"vendor '{vendor}' has no question map yet (needs questionnaire decode) — aborting.")
        return
    stage = f"dakavara_pa.survey_stg_{vendor}_{sid}"
    print(f"file={path}\nvendor={vendor} sid={sid} date={survey_date} mode={'COMMIT' if commit else 'DRY-RUN'}\n", flush=True)

    t0 = time.time()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = max(wb.worksheets, key=lambda s: (s.max_column or 0) * (s.max_row or 0))
    hdr = header_index(ws)
    # apply column aliases: point the canonical name at whatever this wave called it
    for alt, primary in prof.get("col_aliases", {}).items():
        if alt in hdr and primary not in hdr:
            hdr[primary] = hdr[alt]
    miss_cols = [c for c in ([prof["ac_col"]] + list(prof["questions"])) if c not in hdr]
    if miss_cols:
        print(f"WARNING: columns not found in '{ws.title}': {miss_cols}\n  available: {list(hdr)[:30]}")

    with dakavara_session() as db:
        base, ac_by_name = load_codebook(db)
        ac_i = hdr.get(prof["ac_col"]); mob_i = hdr.get(prof.get("mobile_col"))
        id_i = hdr.get(prof.get("id_col")); dist_i = hdr.get(prof.get("district_col"))
        pre = strip_code if prof.get("strip_code") else (lambda v: v)

        def resolve_ac(ac_raw, dist_raw):
            n = _norm(ac_raw)
            if n is None:
                return None, False
            if n in ac_by_name:
                return ac_by_name[n], True
            d = _norm(dist_raw) or ""
            for (an, dk), tgt in AC_DISTRICT_ALIASES.items():
                if an == n and dk in d:
                    return ac_by_name.get(tgt), True
            # Dhruva appends reservation category — ivrs_mobiles stores names bare
            n2 = re.sub(r"\s*\((sc|st)\)\s*$", "", n)
            if n2 != n and n2 in ac_by_name:
                return ac_by_name[n2], True
            if n in AC_ALIASES:
                return ac_by_name.get(AC_ALIASES[n]), True
            if n2 in AC_ALIASES:
                return ac_by_name.get(AC_ALIASES[n2]), True
            return None, False

        rows = []
        opt_dist, ignored = Counter(), Counter()
        unmapped = defaultdict(Counter)
        unmatched_ac = Counter()
        respondents = 0
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r == 0:
                continue
            respondents += 1
            ac_raw = pre(row[ac_i]) if ac_i is not None else None
            if prof["ac_kind"] == "name":
                cid, ok = resolve_ac(ac_raw, row[dist_i] if dist_i is not None else None)
                if not ok and ac_raw not in (None, ""):
                    unmatched_ac[str(ac_raw)] += 1
            else:
                cid = int(ac_raw) if ac_raw not in (None, "") else None
            mobile = str(row[mob_i]).strip() if (mob_i is not None and row[mob_i] not in (None, "")) else None
            clip = str(row[id_i]) if (id_i is not None and row[id_i] is not None) else None

            for col, (qid, scope) in prof["questions"].items():
                ci = hdr.get(col)
                if ci is None:
                    continue
                val = pre(row[ci])
                oid, status = resolve_option(val, scope, base)
                if status == "miss":
                    unmapped[col][str(val)] += 1
                    continue
                if status == "ignore":
                    if _norm(row[ci]):
                        ignored[col] += 1
                    continue
                opt_dist[oid] += 1
                rows.append({
                    "mobile_no": mobile, "constituency_id": cid, "ivrs_survey_id": sid,
                    "round_id": 1, "clip_no": clip, "option_no": None,
                    "ivrs_question_id": qid, "ivrs_option_id": oid, "survey_date": survey_date,
                })

        print(f"parsed {len(rows):,} answer rows from {respondents:,} respondents in {(time.time()-t0)/60:.1f} min", flush=True)
        print(f"  option-id dist (top 15): {dict(opt_dist.most_common(15))}")
        print(f"  ignored non-responses:   {dict(ignored)}")
        if unmatched_ac:
            print(f"  ⚠ unmatched AC names ({len(unmatched_ac)}): {dict(unmatched_ac.most_common(8))}")
        if unmapped:
            print("  ⚠ UNMAPPED values (extend SCOPES / SENTINELS):")
            for col, vals in unmapped.items():
                print(f"      [{col}] {dict(vals.most_common(8))}")
        if not unmatched_ac and not unmapped:
            print("  ✓ clean: every AC resolved, every value mapped or ignored.")

        if not commit:
            print("\nDRY-RUN — nothing written. Re-run with --commit to write the staging table.")
            return

        db.execute(text(f"""CREATE TABLE IF NOT EXISTS {stage} (
            mobile_no VARCHAR(15), constituency_id INT, ivrs_survey_id INT, round_id INT,
            clip_no VARCHAR(64), option_no INT, ivrs_question_id INT, ivrs_option_id INT,
            survey_date DATE) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"""))
        db.execute(text(f"TRUNCATE TABLE {stage}")); db.commit()
        ins = text(f"""INSERT INTO {stage}
            (mobile_no,constituency_id,ivrs_survey_id,round_id,clip_no,option_no,ivrs_question_id,ivrs_option_id,survey_date)
            VALUES (:mobile_no,:constituency_id,:ivrs_survey_id,:round_id,:clip_no,:option_no,:ivrs_question_id,:ivrs_option_id,:survey_date)""")
        t1 = time.time(); n = 0
        for j in range(0, len(rows), batch):
            db.execute(ins, rows[j:j + batch]); db.commit(); n += len(rows[j:j + batch])
            if n % 100000 < batch:
                print(f"  staged {n:,}/{len(rows):,}", flush=True)
        print(f"\nSTAGED {n:,} rows into {stage} in {(time.time()-t1)/60:.1f} min.")
        print(f"Next: review, then merge into survey24.ivrs_survey_answer (adapt scripts/merge_ivrs.py STAGE={stage}, sid={sid}).")


if __name__ == "__main__":
    main()
