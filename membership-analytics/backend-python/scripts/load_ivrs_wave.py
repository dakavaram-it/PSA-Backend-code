"""Load a new IVRS wave Excel (e.g. IVRS_04-06-2026.xlsx) into survey24.ivrs_survey_answer.

Two phases (decoupled — reading and writing never interleave, which keeps the DB
connection healthy):
  1) read all data rows from the sheets (openpyxl, streaming) into memory + validate
  2) in a single DB session: DELETE existing rows for the SID, then INSERT in
     committed batches.

Sheets: 'Details' (definitions, ignored) + data sheets 'SHEET_1','SHEET_2',… with:
  MOBILE NO | OPTION | ACID(constituency_id) | ASSEMBLY | MEMBER TYPE |
  ivrs_survey_id | round_id | clip_no | option_no | ivrs_question_id | ivrs_option_id | Query

Usage:
    PYTHONPATH=. python3.13 scripts/load_ivrs_wave.py "<file.xlsx>"            # dry-run
    PYTHONPATH=. python3.13 scripts/load_ivrs_wave.py "<file.xlsx>" --commit    # delete SID + insert
Options: --sid N (default 31)  --date YYYY-MM-DD (default 2026-06-04)  --batch N (default 5000)
"""
import sys
import time
from collections import Counter

import openpyxl
from sqlalchemy import text

from app.database.db import dakavara_session

COL = {"mobile": 0, "acid": 2, "sid": 5, "round": 6, "clip": 7,
       "option_no": 8, "qid": 9, "opt_id": 10, "member": 4}
DATA_SHEETS = ("SHEET_1", "SHEET_2")
INS = text("""
    INSERT INTO survey24.ivrs_survey_answer
      (mobile_no, constituency_id, ivrs_survey_id, round_id, clip_no, option_no,
       ivrs_question_id, ivrs_option_id, is_deleted, survey_date)
    VALUES (:mobile_no, :constituency_id, :ivrs_survey_id, :round_id, :clip_no, :option_no,
            :ivrs_question_id, :ivrs_option_id, 'N', :survey_date)
""")


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def _int(v):
    return int(v) if v not in (None, "") else None


def main():
    path = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
    if not path:
        print("usage: load_ivrs_wave.py <xlsx> [--commit] [--sid N] [--date YYYY-MM-DD]"); return
    commit = "--commit" in sys.argv
    sid = int(arg("--sid", "31"))
    survey_date = arg("--date", "2026-06-04")
    batch_size = int(arg("--batch", "5000"))
    print(f"file={path}\nmode={'COMMIT' if commit else 'DRY-RUN'} sid={sid} date={survey_date} batch={batch_size}\n", flush=True)

    # ── phase 1: read + validate ──
    t0 = time.time()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in DATA_SHEETS if s in wb.sheetnames] or [s for s in wb.sheetnames if s.lower().startswith("sheet")]
    opt_dist, member_dist, sid_dist = Counter(), Counter(), Counter()
    rows, seen, skipped = [], 0, 0
    for sn in sheets:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            seen += 1
            r_sid = _int(row[COL["sid"]]); opt_id = _int(row[COL["opt_id"]]); mobile = row[COL["mobile"]]
            sid_dist[r_sid] += 1
            if r_sid != sid or opt_id is None or mobile in (None, ""):
                skipped += 1; continue
            opt_dist[opt_id] += 1
            member_dist[str(row[COL["member"]])] += 1
            rows.append({
                "mobile_no": str(mobile).strip(), "constituency_id": _int(row[COL["acid"]]),
                "ivrs_survey_id": sid, "round_id": _int(row[COL["round"]]),
                "clip_no": (str(row[COL["clip"]]) if row[COL["clip"]] is not None else None),
                "option_no": _int(row[COL["option_no"]]), "ivrs_question_id": _int(row[COL["qid"]]),
                "ivrs_option_id": opt_id, "survey_date": survey_date,
            })
    print(f"read {seen:,} rows ({len(rows):,} valid sid={sid}, {skipped:,} skipped) in {(time.time()-t0)/60:.1f} min", flush=True)
    print(f"  survey-id dist: {dict(sid_dist)}")
    print(f"  option-id dist: {dict(opt_dist)}  (28=TDP/NDA, 15=YSRCP, 16=Others)")
    print(f"  member dist:    {dict(member_dist)}")

    if not commit:
        print("\nDRY-RUN only — re-run with --commit to insert."); return

    # ── phase 2: delete existing SID, then batched insert (resilient to the
    #    intermittent metadata lock on this shared table) ──
    from sqlalchemy.exc import OperationalError

    def run_resilient(db, stmt, params=None, what="stmt"):
        """Execute+commit; on lock timeout (1205) or deadlock (1213) retry forever
        with short backoff, so we slip through whenever the MDL window opens."""
        waits = 0
        while True:
            try:
                db.execute(stmt, params) if params is not None else db.execute(stmt)
                db.commit(); return waits
            except OperationalError as e:
                db.rollback()
                code = e.orig.args[0] if e.orig and e.orig.args else None
                if code in (1205, 1213):
                    waits += 1
                    if waits % 6 == 1:
                        print(f"  [{what}] table busy (lock), retrying…", flush=True)
                    time.sleep(5); continue
                raise

    t1 = time.time()
    inserted = 0
    with dakavara_session() as db:
        db.execute(text("SET SESSION lock_wait_timeout=10"))          # metadata-lock timeout (fail fast)
        db.execute(text("SET SESSION innodb_lock_wait_timeout=10"))   # row-lock timeout
        existing = db.execute(text("SELECT COUNT(*) FROM survey24.ivrs_survey_answer WHERE ivrs_survey_id=:s"),
                              {"s": sid}).scalar()
        if existing:
            print(f"deleting {existing:,} existing rows for sid={sid}…", flush=True)
            run_resilient(db, text("DELETE FROM survey24.ivrs_survey_answer WHERE ivrs_survey_id=:s"), {"s": sid}, "delete")
        for j in range(0, len(rows), batch_size):
            run_resilient(db, INS, rows[j:j + batch_size], f"batch@{j}")
            inserted += len(rows[j:j + batch_size])
            if inserted % 100000 < batch_size:
                print(f"  inserted {inserted:,}/{len(rows):,} ({(time.time()-t1)/60:.1f} min)", flush=True)
    print(f"\nINSERTED {inserted:,} rows into survey24.ivrs_survey_answer in {(time.time()-t1)/60:.1f} min "
          f"(total {(time.time()-t0)/60:.1f} min)", flush=True)


if __name__ == "__main__":
    main()
